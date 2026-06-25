"""
Purpose:
    Build MMB IRF and regression-format datasets from legacy source exports.

Inputs:
    ../input/responses/*.csv
    ../input/sacratios_csv/*.csv
    ../input/Model_Characteristics_corrections_llm.xlsx
    ../input/bob_var_irfs.csv
    ../../../../config/params.yaml

Outputs:
    ../output/MMB_IRF_format.dta
    ../output/MMB_IRF_format.csv
    ../output/MMB_IRF_format_codebook.txt
    ../output/MMB_IRF_format_full.dta
    ../output/MMB_IRF_format_full_codebook.txt
    ../output/MMB_reg_format.dta
    ../output/MMB_reg_format.xlsx
    ../output/MMB_reg_format_codebook.txt

Run:
    make from tasks/data/build_mmb_analysis_dataset/code/
"""

from pathlib import Path
import math
import re

import numpy as np
from openpyxl import load_workbook
import pandas as pd
import yaml


code_dir = Path(__file__).resolve().parent
task_dir = code_dir.parent
repo_dir = task_dir.parents[2]
input_dir = task_dir / "input"
output_dir = task_dir / "output"
config_path = repo_dir / "config" / "params.yaml"
stata_float_min = float(np.float32(2 ** -127))


# Keep research parameters in standard YAML and read the MMB block directly.
with open(config_path, "r", encoding="utf-8") as f:
    params = yaml.safe_load(f)["mmb"]

qforgive = int(params["qforgive"])
horizons = [int(x) for x in params["horizons"]]
drop_model_rule = int(params["drop_model_rule"])
pi_in_sacratio = int(params["pi_in_sacratio"])
cloud_graph_extent = int(params["cloud_graph_extent"])
citation_weight_as_of = str(params["citation_weight_as_of"])
duplicate_models_to_drop = set(params["duplicate_models_to_drop"])
excluded_models_to_drop = set(params["excluded_models_to_drop"])
sacrifice_ratio_models_to_drop = set(params["sacrifice_ratio_models_to_drop"])
output_response_labels = set(params["output_response_labels"])
model_name_replacements = dict(params.get("model_name_replacements", {}))

output_dir.mkdir(parents=True, exist_ok=True)


def clean_bool(value):
    if pd.isna(value):
        return np.nan
    if isinstance(value, (bool, np.bool_)):
        return int(value)
    if isinstance(value, (int, float, np.integer, np.floating)):
        return int(value)
    text = str(value).strip().lower()
    if text in ["true", "yes", "1"]:
        return 1
    if text in ["false", "no", "0", "no mention"]:
        return 0
    return np.nan


def quarter_string(value):
    if pd.isna(value) or value == "":
        return "."
    text = str(value).strip()
    if "q" in text.lower() and len(text) >= 6:
        if ":" in text:
            return "."
        year, quarter = text.lower().replace(" ", "").replace(":", "").split("q", 1)
        roman_quarters = {"i": 1, "ii": 2, "iii": 3, "iv": 4}
        quarter_value = roman_quarters.get(quarter, quarter)
        return f"{int(year)}q{int(quarter_value)}"
    date_value = pd.to_datetime(value)
    quarter = math.ceil(date_value.month / 3)
    return f"{date_value.year}q{quarter}"


def quarter_decimal_year(value):
    text = str(value)
    if not text or text == ".":
        return np.nan
    if "q" not in text.lower():
        return np.nan
    year, quarter = text.lower().split("q", 1)
    return float(year) + (float(quarter) - 1.0) / 4.0


def write_codebook(df, path, source_note, filters_note):
    with open(path, "w", encoding="utf-8") as f:
        f.write(f"{path.stem.replace('_codebook', '')} codebook\n")
        f.write("=" * (len(path.stem.replace("_codebook", "")) + 9))
        f.write("\n\n")
        f.write(f"Rows: {df.shape[0]}\n")
        f.write(f"Columns: {df.shape[1]}\n")
        f.write(f"Source: {source_note}\n")
        f.write(f"Filters and transformations: {filters_note}\n")
        if "period" in df.columns:
            f.write(f"Period coverage: {df['period'].min()} to {df['period'].max()}\n")
        if "model" in df.columns:
            f.write(f"Models: {df['model'].nunique()}\n")
        if "rule" in df.columns:
            f.write(f"Rules: {', '.join(sorted(df['rule'].dropna().astype(str).unique()))}\n")
        f.write("\nColumns and dtypes\n")
        f.write("------------------\n")
        for name, dtype in df.dtypes.items():
            f.write(f"{name}: {dtype}\n")
        f.write("\nFirst 10 rows\n")
        f.write("-------------\n")
        f.write(df.head(10).to_string(index=False))
        f.write("\n")


def compare_for_report(name, rebuilt, legacy_path, key_columns):
    lines = [name, "-" * len(name)]
    if not legacy_path.exists():
        lines.append(f"legacy file missing: {legacy_path}")
        return lines

    try:
        legacy = pd.read_stata(legacy_path, convert_categoricals=False)
    except Exception as exc:
        lines.append(f"legacy file unreadable: {legacy_path.relative_to(repo_dir)}")
        lines.append(f"read error: {type(exc).__name__}: {exc}")
        return lines
    lines.append(f"legacy: {legacy_path.relative_to(repo_dir)}")
    lines.append(f"rebuilt shape: {rebuilt.shape}")
    lines.append(f"legacy shape: {legacy.shape}")

    rebuilt_columns = list(rebuilt.columns)
    legacy_columns = list(legacy.columns)
    added_columns = [c for c in rebuilt_columns if c not in legacy_columns]
    dropped_columns = [c for c in legacy_columns if c not in rebuilt_columns]
    common_columns = [c for c in rebuilt_columns if c in legacy_columns]
    lines.append(f"columns added: {added_columns if added_columns else 'none'}")
    lines.append(f"columns dropped: {dropped_columns if dropped_columns else 'none'}")

    if not all(c in rebuilt.columns and c in legacy.columns for c in key_columns):
        lines.append("row comparison skipped because key columns are missing")
        return lines

    rebuilt_keyed = rebuilt[common_columns].copy()
    legacy_keyed = legacy[common_columns].copy()
    rebuilt_keyed["_rebuilt_row"] = 1
    legacy_keyed["_legacy_row"] = 1
    merged = rebuilt_keyed.merge(
        legacy_keyed,
        on=key_columns,
        how="outer",
        suffixes=("_rebuilt", "_legacy"),
        indicator="__report_merge",
    )
    added_rows = merged["__report_merge"].eq("left_only").sum()
    dropped_rows = merged["__report_merge"].eq("right_only").sum()
    lines.append(f"rows added: {int(added_rows)}")
    lines.append(f"rows dropped: {int(dropped_rows)}")

    matched = merged.loc[merged["__report_merge"].eq("both")].copy()
    changed_columns = []
    numeric_max_diffs = {}
    changed_models = set()
    compare_columns = [c for c in common_columns if c not in key_columns]
    for col in compare_columns:
        left = matched[f"{col}_rebuilt"]
        right = matched[f"{col}_legacy"]
        left_num = pd.to_numeric(left, errors="coerce")
        right_num = pd.to_numeric(right, errors="coerce")
        numeric_col = left_num.notna().any() or right_num.notna().any()
        if numeric_col:
            diff = (left_num - right_num).abs()
            max_diff = diff.max(skipna=True)
            different = diff.gt(1e-6) | (left_num.isna() != right_num.isna())
            if pd.notna(max_diff):
                numeric_max_diffs[col] = float(max_diff)
        else:
            different = left.fillna("<missing>").astype(str) != right.fillna("<missing>").astype(str)
        if different.any():
            changed_columns.append(col)
            if "model" in matched.columns:
                changed_models.update(matched.loc[different, "model"].dropna().astype(str))

    lines.append(f"columns changed: {changed_columns if changed_columns else 'none'}")
    if numeric_max_diffs:
        max_col = max(numeric_max_diffs, key=numeric_max_diffs.get)
        lines.append(f"max numeric difference for unchanged/common columns: {max_col} = {numeric_max_diffs[max_col]:.8g}")
    else:
        lines.append("max numeric difference for unchanged/common columns: none")
    if changed_models:
        listed_models = sorted(changed_models)
        suffix = "" if len(listed_models) <= 25 else f" ... ({len(listed_models)} total)"
        lines.append(f"models with changed values: {', '.join(listed_models[:25])}{suffix}")
    else:
        lines.append("models with changed values: none")

    exact_pass = (
        rebuilt.shape == legacy.shape
        and not added_columns
        and not dropped_columns
        and int(added_rows) == 0
        and int(dropped_rows) == 0
        and not changed_columns
    )
    lines.append(f"status: {'PASS' if exact_pass else 'DIFF'}")
    return lines


# Step 1: Import all model-rule IRFs and convert them to a long panel.
response_frames = []
for csv_path in sorted((input_dir / "responses").glob("*.csv")):
    raw = pd.read_csv(csv_path)
    period_columns = [c for c in raw.columns if str(c).isdigit()]
    long = raw.melt(
        id_vars=["resulttype", "rule", "model", "shock", "variable"],
        value_vars=period_columns,
        var_name="period",
        value_name="response",
    )
    response_frames.append(long)

responses = pd.concat(response_frames, ignore_index=True)
responses["model"] = responses["model"].replace(model_name_replacements)
responses["period"] = responses["period"].astype(int)
responses["response"] = pd.to_numeric(responses["response"], errors="coerce")

responses["model_type"] = responses["model"].str.slice(0, 2)
responses.loc[responses["model_type"] == "RB", "model_type"] = "NK"
responses.loc[responses["model_type"].isin(["G2", "G3", "G7"]), "model_type"] = "Other"
responses.loc[responses["model_type"] == "NK", "model_type"] = "Calibrated"
responses.loc[responses["model_type"] == "US", "model_type"] = "Estimated"

responses["variable_clean"] = responses["variable"].replace(
    {
        "Inflation": "pi",
        "inflationq": "piq",
        "Interest": "irate",
    }
)
responses.loc[responses["variable"].isin(output_response_labels), "variable_clean"] = "y"

drop_models = duplicate_models_to_drop | excluded_models_to_drop
responses = responses.loc[~responses["model"].isin(drop_models)].copy()
responses = responses.loc[responses["variable_clean"].isin(["pi", "piq", "irate", "y"])].copy()

# Flip signs so responses are to an expansionary monetary policy shock.
# Stata stored these imported values at float precision; matching that rounding
# keeps the Python output aligned with the legacy datasets.
responses["response"] = (-responses["response"]).astype(np.float32).astype(float)
responses.loc[responses["response"].abs() < stata_float_min, "response"] = 0.0
responses["period"] = responses["period"] - 1
responses = responses.loc[responses["period"] != 0].copy()
responses["rule"] = responses["rule"].replace({"Inertial Taylor": "Inertial_Taylor"})

irf_wide = (
    responses.pivot_table(
        index=["model", "rule", "period"],
        columns="variable_clean",
        values="response",
        aggfunc="mean",
    )
    .reset_index()
    .rename_axis(None, axis=1)
)
for col in ["piq", "y", "irate", "pi"]:
    if col not in irf_wide.columns:
        irf_wide[col] = np.nan
    irf_wide[col] = irf_wide[col].astype(np.float32).astype(float)

irf_wide = irf_wide.sort_values(["model", "rule", "period"]).reset_index(drop=True)
irf_wide["id"] = irf_wide.groupby(["model", "rule"], sort=True).ngroup() + 1
irf_wide = irf_wide.loc[irf_wide["rule"] != "Model"].copy()
rrate_next_piq = irf_wide.groupby(["model", "rule"])["piq"].shift(-1).astype(np.float32)
irf_wide["rrate"] = (irf_wide["irate"].astype(np.float32) - rrate_next_piq).astype(np.float32).astype(float)
irf_wide.loc[irf_wide["rrate"].abs() < stata_float_min, "rrate"] = 0.0
irf = irf_wide[["id", "period", "model", "rule", "piq", "y", "irate", "rrate", "pi"]].copy()

irf.to_stata(output_dir / "MMB_IRF_format.dta", write_index=False, version=118)
irf.to_csv(output_dir / "MMB_IRF_format.csv", index=False)


# Step 2: Build model-level attributes from the hand-coded workbook after LLM corrections.
model_characteristics_path = input_dir / "Model_Characteristics_corrections_llm.xlsx"
rhs_raw = pd.read_excel(model_characteristics_path, engine="openpyxl")
rhs_raw.columns = [c.lower() for c in rhs_raw.columns]

# Some cb_authors cells are formulas like =1/2. If the workbook lacks cached
# formula results, read those simple fractions from the formula text itself.
if rhs_raw["cb_authors"].isna().any():
    wb = load_workbook(model_characteristics_path, data_only=False, read_only=True)
    ws = wb.active
    headers = [str(c.value).lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    model_col = headers.index("model")
    cb_authors_col = headers.index("cb_authors")
    formula_values = {}
    for row in ws.iter_rows(min_row=2, max_row=93):
        model = row[model_col].value
        value = row[cb_authors_col].value
        if isinstance(value, str) and value.startswith("="):
            formula = value[1:].strip()
            match = re.fullmatch(r"([0-9.]+)\s*/\s*([0-9.]+)", formula)
            if match:
                numerator = float(match.group(1))
                denominator = float(match.group(2))
                formula_values[model] = numerator / denominator
            elif re.fullmatch(r"[0-9.]+", formula):
                formula_values[model] = float(formula)
    rhs_raw.loc[rhs_raw["cb_authors"].isna(), "cb_authors"] = (
        rhs_raw.loc[rhs_raw["cb_authors"].isna(), "model"].map(formula_values)
    )

rhs = rhs_raw.iloc[:92].copy()

keep_columns = [
    "model",
    "date_pub",
    "cb_authors",
    "est_date_range_start",
    "est_date_range_end",
    "estimated",
    "calibrated",
    "price_indexation",
    "open",
    "firm_balance_sheet_channel",
    "bank_lending_intermediation_channel",
    "constrained_household_demand_channel",
    "real_labor_market_friction",
    "tax",
    "gov_spend",
    "gov_debt",
    "sticky_price_method",
    "sticky_wages",
    "wage_indexation",
    "wage_index_method",
    "final_mmb_count",
    "learning",
    "cites",
]
rhs = rhs[keep_columns].copy()

for col in ["estimated", "calibrated", "open", "tax", "gov_spend", "gov_debt", "learning"]:
    rhs[col] = rhs[col].map(clean_bool)

rhs.loc[rhs["model"].isin(["US_VMDno", "US_VMDop", "US_FGKR15", "US_IAC05"]), "calibrated"] = 0
rhs.loc[rhs["model"].isin(["US_IAC05", "US_FGKR15"]), "estimated"] = 1
rhs.loc[rhs["model"] == "US_VMDno", "calibrated"] = 1

rhs["pub_date"] = rhs["date_pub"].map(quarter_string)
rhs["est_start"] = rhs["est_date_range_start"].map(quarter_string)
rhs["est_end"] = rhs["est_date_range_end"].map(quarter_string)
rhs = rhs.drop(columns=["date_pub", "est_date_range_start", "est_date_range_end"])

pub_year = rhs["pub_date"].str.extract(r"^(\d{4})")[0].astype(float)
est_year = rhs["est_start"].str.extract(r"^(\d{4})")[0].astype(float)
rhs["est_early"] = np.where(est_year.notna(), (est_year < 1980).astype(int), np.nan)
rhs["est_late"] = np.where(est_year.notna(), (est_year >= 1980).astype(int), np.nan)
rhs["vint_early"] = np.where(pub_year.notna(), (pub_year < 2000).astype(int), np.nan)
rhs["vint_mid"] = np.where(pub_year.notna(), ((pub_year >= 2000) & (pub_year <= 2007)).astype(int), np.nan)
rhs["vint_late"] = np.where(pub_year.notna(), (pub_year > 2007).astype(int), np.nan)

sticky_price_method = rhs["sticky_price_method"].fillna("")
rhs["stky_pr_calvo"] = sticky_price_method.isin(["Calvo", "Calvo-like"]).astype(int)
rhs["stky_pr_rotemberg"] = (sticky_price_method == "Rotemberg").astype(int)
rhs["stky_pr_other"] = (~sticky_price_method.isin(["Calvo", "Calvo-like", "Rotemberg", "NA"])).astype(int)
rhs["not_stky_pr"] = (sticky_price_method == "NA").astype(int)
rhs["stky_pr"] = (rhs["not_stky_pr"] == 0).astype(int)
rhs = rhs.drop(columns=["sticky_price_method"])

rhs["stky_wg"] = rhs["sticky_wages"].map(clean_bool)
rhs["pr_ndx"] = rhs["price_indexation"].map(clean_bool)
rhs["wg_ndx"] = rhs["wage_indexation"].map(clean_bool)
rhs = rhs.drop(columns=["sticky_wages", "price_indexation", "wage_indexation"])

wage_index_method = rhs["wage_index_method"].fillna("")
rhs["wg_ndx_mult"] = (wage_index_method == "Multiple").astype(int)
rhs["wg_ndx_prprice"] = wage_index_method.isin(["Prev Price Inflation", "Prev Agg Inflation"]).astype(int)
rhs["wg_ndx_other"] = wage_index_method.isin(
    ["Other", "Prev Wage Inflation", "Prev Wages", "Steady-State Inflation"]
).astype(int)
rhs = rhs.drop(columns=["wage_index_method"])

rhs["stky_pr_ndx"] = ((rhs["stky_pr"] == 1) & (rhs["pr_ndx"] == 1)).astype(int)
rhs["stky_pr_nondx"] = ((rhs["stky_pr"] == 1) & (rhs["pr_ndx"] == 0)).astype(int)
rhs["stky_wg_ndx"] = ((rhs["stky_wg"] == 1) & (rhs["wg_ndx"] == 1)).astype(int)
rhs["stky_wg_nondx"] = ((rhs["stky_wg"] == 1) & (rhs["wg_ndx"] == 0)).astype(int)
rhs["stky_all"] = ((rhs["stky_pr"] == 1) & (rhs["stky_wg"] == 1)).astype(int)
rhs["stky_pr_wg_ndx"] = ((rhs["stky_pr"] == 1) & (rhs["wg_ndx"] == 1)).astype(int)
rhs["ndx_all"] = ((rhs["pr_ndx"] == 1) & (rhs["wg_ndx"] == 1)).astype(int)

rhs["cites"] = pd.to_numeric(rhs["cites"], errors="coerce")
rhs["cites_w1"] = np.log1p(rhs["cites"])
rhs["cites_w2"] = np.sqrt(1 + rhs["cites"])
rhs["cites_w3"] = np.minimum(rhs["cites"], rhs["cites"].quantile(0.95))

# Age-adjust citation counts as of the configured month, then normalize the log weight.
as_of_period = pd.Period(citation_weight_as_of, freq="M")
as_of_decimal_year = as_of_period.year + (as_of_period.month - 1) / 12
pub_decimal_year = rhs["pub_date"].map(quarter_decimal_year)
rhs["citation_years_since_pub"] = as_of_decimal_year - pub_decimal_year
rhs.loc[rhs["citation_years_since_pub"] < 0, "citation_years_since_pub"] = np.nan
rhs["cites_per_year"] = (1 + rhs["cites"]) / (1 + rhs["citation_years_since_pub"])
rhs["citation_weight_raw"] = np.log1p(rhs["cites_per_year"])
rhs["citation_weight"] = rhs["citation_weight_raw"] / rhs["citation_weight_raw"].mean()

channel_columns = [
    "firm_balance_sheet_channel",
    "bank_lending_intermediation_channel",
    "constrained_household_demand_channel",
    "real_labor_market_friction",
]
for col in channel_columns:
    rhs[col] = rhs[col].map(clean_bool)
rhs["other_channel"] = (
    (rhs["firm_balance_sheet_channel"] == 1)
    | (rhs["bank_lending_intermediation_channel"] == 1)
    | (rhs["constrained_household_demand_channel"] == 1)
    | (rhs["real_labor_market_friction"] == 1)
).astype(int)

rhs["fiscal"] = ((rhs["tax"] == 1) | (rhs["gov_debt"] == 1) | (rhs["gov_spend"] == 1)).astype(int)
rhs = rhs.rename(
    columns={
        "final_mmb_count": "neq",
        "firm_balance_sheet_channel": "firm_bs",
        "bank_lending_intermediation_channel": "bank",
        "constrained_household_demand_channel": "hh_demand",
        "real_labor_market_friction": "labor_frict",
    }
)
rhs["ln_neq"] = np.log(pd.to_numeric(rhs["neq"], errors="coerce"))
cb_authors_numeric = pd.to_numeric(rhs["cb_authors"], errors="coerce")
rhs["cb_authors_ext"] = np.where(cb_authors_numeric.notna(), (cb_authors_numeric != 0).astype(int), np.nan)





#MANUAL OVERRIDES THAT CONNOR PUT IN
rhs.loc[rhs['model'] == 'US_YR13', 'learning'] = 0
# rhs.loc[rhs["model"] == "US_FGKR15", "estimated"] = 1
# rhs.loc[rhs["model"] == "US_VMDno", "calibrated"] = 1





rhs.to_stata(output_dir / "rhs.dta", write_index=False, version=118)


# Step 3: Construct outcome variables for each horizon.
lhs_by_horizon = {}
for horizon in horizons:
    d = irf.sort_values(["id", "period"]).copy()

    for var in ["piq", "y", "irate", "rrate", "pi"]:
        previous = d.groupby("id")[var].shift(1)
        changed = (np.sign(previous) != np.sign(d[var])) & (d["period"] > qforgive)
        changed = changed & previous.notna() & d[var].notna()
        d[f"{var}_chg_sign"] = changed.astype(int)

        d[f"{var}_chg_sign_sum"] = 0
        d.loc[d["period"] >= qforgive + 2, f"{var}_chg_sign_sum"] = (
            d.loc[d["period"] >= qforgive + 2].groupby("id")[f"{var}_chg_sign"].cumsum()
        )

    d["flag_piq_wrongsign"] = ((d["piq"] < 0) & (d["period"] == qforgive + 1)).astype(int)
    d["flag_y_wrongsign"] = ((d["y"] < 0) & (d["period"] == qforgive + 1)).astype(int)
    d["flag_pi_wrongsign"] = ((d["pi"] < 0) & (d["period"] == qforgive + 1)).astype(int)

    rows = []
    for group_id, g in d.groupby("id", sort=True):
        row = {
            "id": group_id,
            "model": g["model"].iloc[0],
            "rule": g["rule"].iloc[0],
        }
        horizon_rows = g.loc[g["period"] <= horizon]

        for var in ["piq", "y", "irate", "rrate", "pi"]:
            row[f"{var}_cum{horizon}"] = horizon_rows[var].sum(skipna=False)
            if var in ["irate", "rrate"]:
                shock_rows = g.loc[g["period"] == 1, var]
                row[f"{var}_shock"] = shock_rows.iloc[0] if len(shock_rows) else np.nan

            row[f"{var}_value_min"] = g[var].min(skipna=True)
            min_periods = g.loc[g[var] == row[f"{var}_value_min"], "period"]
            row[f"{var}_timing_min"] = min_periods.mean() if len(min_periods) else np.nan

            row[f"{var}_value_max"] = g[var].max(skipna=True)
            max_periods = g.loc[g[var] == row[f"{var}_value_max"], "period"]
            row[f"{var}_timing_max"] = max_periods.mean() if len(max_periods) else np.nan

            row[f"{var}_integral"] = g[var].sum(skipna=True)
            peak_abs = g[var].abs().max(skipna=True)
            peak_rows = g.loc[g[var].abs() == peak_abs]
            row[f"{var}_value_peak"] = peak_rows[var].mean() if len(peak_rows) else np.nan
            row[f"{var}_timing_peak"] = peak_rows["period"].mean() if len(peak_rows) else np.nan

            row[f"{var}_chg_sign"] = g[f"{var}_chg_sign"].sum()

        row["flag_piq_wrongsign"] = g["flag_piq_wrongsign"].sum()
        row["flag_y_wrongsign"] = g["flag_y_wrongsign"].sum()
        row["flag_pi_wrongsign"] = g["flag_pi_wrongsign"].sum()
        rows.append(row)

    lhs = pd.DataFrame(rows)

    for var in ["piq", "y", "irate", "rrate", "pi"]:
        lhs[f"{var}_timing_theorypeak"] = lhs[f"{var}_timing_max"]
        lhs[f"{var}_value_theorypeak"] = lhs[f"{var}_value_max"]
        lhs[f"{var}_timing_agnpeak"] = np.where(
            lhs[f"{var}_integral"] > 0,
            lhs[f"{var}_timing_max"],
            np.where(lhs[f"{var}_integral"] < 0, lhs[f"{var}_timing_min"], np.nan),
        )
        lhs[f"{var}_value_agnpeak"] = np.where(
            lhs[f"{var}_integral"] > 0,
            lhs[f"{var}_value_max"],
            np.where(lhs[f"{var}_integral"] < 0, lhs[f"{var}_value_min"], np.nan),
        )
        lhs = lhs.drop(columns=[f"{var}_integral"])

    lhs[f"IScurve{horizon}"] = lhs[f"y_cum{horizon}"] / lhs[f"rrate_cum{horizon}"]
    lhs[f"infl_per_rr{horizon}"] = lhs[f"piq_cum{horizon}"] / lhs[f"rrate_cum{horizon}"]
    lhs[f"Billsacrat{horizon}"] = lhs[f"piq_cum{horizon}"] / lhs[f"y_cum{horizon}"]

    lhs["rule_tr"] = (lhs["rule"] == "Taylor").astype(int)
    lhs["rule_itr"] = (lhs["rule"] == "Inertial_Taylor").astype(int)
    lhs["rule_g"] = (lhs["rule"] == "Growth").astype(int)
    lhs["rule_m"] = (lhs["rule"] == "Model").astype(int)
    if drop_model_rule == 1:
        lhs = lhs.loc[lhs["rule_m"] != 1].drop(columns=["rule_m"])

    lhs_by_horizon[horizon] = lhs.copy()
    lhs.to_stata(output_dir / f"lhs_{horizon}.dta", write_index=False, version=118)


# Step 4: Compute sacrifice ratios from inflation-target shock responses.
sacratio_by_horizon = {}
for horizon in horizons:
    sac_frames = []
    for csv_path in sorted((input_dir / "sacratios_csv").glob("*.csv")):
        raw = pd.read_csv(csv_path)
        period_columns = [c for c in raw.columns if str(c).isdigit() and int(c) <= horizon + 1]
        long = raw.melt(
            id_vars=["resulttype", "rule", "model", "shock", "variable"],
            value_vars=period_columns,
            var_name="period",
            value_name="response",
        )
        sac_frames.append(long)

    sac = pd.concat(sac_frames, ignore_index=True)
    sac["period"] = sac["period"].astype(int) - 1
    sac = sac.loc[sac["period"] != 0].copy()
    sac["response"] = pd.to_numeric(sac["response"], errors="coerce").astype(np.float32)
    sac.loc[sac["response"].abs() < stata_float_min, "response"] = 0.0
    sac = sac.loc[~sac["model"].isin(sacrifice_ratio_models_to_drop)].copy()
    sac["rule"] = sac["rule"].replace(
        {
            "inertial": "Inertial_Taylor",
            "taylor": "Taylor",
            "growth": "Growth",
            "model": "Model",
        }
    )
    if drop_model_rule == 1:
        sac = sac.loc[sac["rule"] != "Model"].copy()

    sac["variable_clean"] = sac["variable"].replace({"Inflation": "pi", "inflationq": "piq"})
    sac.loc[sac["variable"].isin(output_response_labels), "variable_clean"] = "ygap"
    sac = sac.sort_values(["model", "rule", "variable_clean", "period"]).copy()
    sac_rows = []
    for (model, rule), group in sac.groupby(["model", "rule"], sort=True):
        y_values = group.loc[
            (group["variable_clean"] == "ygap") & (group["period"] <= horizon),
            "response",
        ].to_numpy(dtype=np.float32)
        y_cum = np.sum(y_values, dtype=np.float32)
        pi_value = group.loc[
            (group["variable_clean"] == "pi") & (group["period"] == horizon),
            "response",
        ]
        piq_value = group.loc[
            (group["variable_clean"] == "piq") & (group["period"] == horizon),
            "response",
        ]
        pi_value = np.float32(pi_value.iloc[0]) if len(pi_value) else np.float32(np.nan)
        piq_value = np.float32(piq_value.iloc[0]) if len(piq_value) else np.float32(np.nan)
        if pi_in_sacratio == 1:
            ratio = np.float32(y_cum / (pi_value + 0.00000000001))
        else:
            ratio = np.float32(y_cum / (piq_value + 0.00000000001))
        sac_rows.append({"model": model[:-6], "rule": rule, f"sacratio{horizon}": ratio})

    sac_out = pd.DataFrame(sac_rows)
    sac_out["model"] = sac_out["model"].replace(model_name_replacements)
    sacratio_by_horizon[horizon] = sac_out
    sac_out.to_stata(output_dir / f"sacratios_{horizon}.dta", write_index=False, version=118)


# Step 5: Merge the cross-sectional regression panel.
reg = lhs_by_horizon[horizons[0]].copy()
for horizon in horizons[1:]:
    new_columns = ["model", "rule"] + [c for c in lhs_by_horizon[horizon].columns if c not in reg.columns]
    reg = reg.merge(lhs_by_horizon[horizon][new_columns], on=["model", "rule"], how="left")

for horizon in horizons:
    reg = reg.merge(sacratio_by_horizon[horizon], on=["model", "rule"], how="left")

reg = reg.merge(rhs, on="model", how="left")
reg = reg.loc[~reg["model"].isin(drop_models)].copy()
if "id" in reg.columns:
    reg = reg.drop(columns=["id"])

legacy_reg_columns = [
    "model", "rule", "rule_tr", "rule_itr", "rule_g",
    "piq_value_min", "y_value_min", "irate_value_min", "rrate_value_min", "pi_value_min",
    "piq_value_max", "y_value_max", "irate_value_max", "rrate_value_max", "pi_value_max",
    "piq_timing_min", "y_timing_min", "irate_timing_min", "rrate_timing_min", "pi_timing_min",
    "piq_timing_max", "y_timing_max", "irate_timing_max", "rrate_timing_max", "pi_timing_max",
    "piq_cum20", "y_cum20", "irate_cum20", "rrate_cum20", "pi_cum20",
    "piq_cum40", "y_cum40", "irate_cum40", "rrate_cum40", "pi_cum40",
    "piq_cum60", "y_cum60", "irate_cum60", "rrate_cum60", "pi_cum60",
    "sacratio20", "sacratio40", "sacratio60",
    "flag_piq_wrongsign", "flag_y_wrongsign", "flag_pi_wrongsign",
    "piq_chg_sign", "y_chg_sign", "irate_chg_sign", "rrate_chg_sign", "pi_chg_sign",
    "irate_shock", "rrate_shock",
    "cb_authors", "cb_authors_ext", "estimated", "calibrated", "neq", "open", "firm_bs", "bank",
    "hh_demand", "labor_frict", "gov_spend", "gov_debt", "tax", "fiscal", "other_channel", "learning",
    "pr_ndx", "wg_ndx", "wg_ndx_mult", "wg_ndx_prprice", "wg_ndx_other",
    "stky_wg", "stky_pr", "stky_pr_other", "stky_pr_rotemberg", "stky_pr_calvo",
    "not_stky_pr", "stky_pr_ndx", "stky_wg_ndx",
    "vint_early", "vint_mid", "vint_late", "est_early", "est_late", "pub_date", "est_start", "est_end",
    "cites", "citation_years_since_pub", "cites_per_year", "citation_weight_raw", "citation_weight",
    "piq_value_peak", "y_value_peak", "irate_value_peak", "rrate_value_peak", "pi_value_peak",
    "piq_timing_peak", "y_timing_peak", "irate_timing_peak", "rrate_timing_peak", "pi_timing_peak",
    "piq_timing_theorypeak", "piq_value_theorypeak", "y_timing_theorypeak", "y_value_theorypeak",
    "irate_timing_theorypeak", "irate_value_theorypeak", "rrate_timing_theorypeak", "rrate_value_theorypeak",
    "pi_timing_theorypeak", "pi_value_theorypeak",
    "piq_timing_agnpeak", "piq_value_agnpeak", "y_timing_agnpeak", "y_value_agnpeak",
    "irate_timing_agnpeak", "irate_value_agnpeak", "rrate_timing_agnpeak", "rrate_value_agnpeak",
    "pi_timing_agnpeak", "pi_value_agnpeak",
    "IScurve20", "infl_per_rr20", "Billsacrat20",
    "IScurve40", "infl_per_rr40", "Billsacrat40",
    "IScurve60", "infl_per_rr60", "Billsacrat60",
    "stky_pr_nondx", "stky_wg_nondx", "stky_all", "stky_pr_wg_ndx", "ndx_all", "ln_neq",
]
reg = reg[legacy_reg_columns].sort_values(["rule", "model"]).reset_index(drop=True)

legacy_reg_float32_columns = [
    "rule_tr", "rule_itr", "rule_g", "piq_value_min", "y_value_min", "irate_value_min",
    "rrate_value_min", "pi_value_min", "piq_value_max", "y_value_max", "irate_value_max",
    "rrate_value_max", "pi_value_max", "piq_timing_min", "y_timing_min", "irate_timing_min",
    "rrate_timing_min", "pi_timing_min", "piq_timing_max", "y_timing_max", "irate_timing_max",
    "rrate_timing_max", "pi_timing_max", "piq_cum20", "y_cum20", "irate_cum20",
    "rrate_cum20", "pi_cum20", "piq_cum40", "y_cum40", "irate_cum40", "rrate_cum40",
    "pi_cum40", "piq_cum60", "y_cum60", "irate_cum60", "rrate_cum60", "pi_cum60",
    "sacratio20", "sacratio40", "sacratio60", "irate_shock", "rrate_shock",
    "cb_authors_ext", "fiscal", "other_channel", "wg_ndx_mult", "wg_ndx_prprice",
    "wg_ndx_other", "stky_pr", "stky_pr_other", "stky_pr_rotemberg", "stky_pr_calvo",
    "not_stky_pr", "stky_pr_ndx", "stky_wg_ndx", "vint_early", "vint_mid", "vint_late",
    "est_early", "est_late", "cites", "citation_years_since_pub", "cites_per_year",
    "citation_weight_raw", "citation_weight", "piq_value_peak", "y_value_peak", "irate_value_peak",
    "rrate_value_peak", "pi_value_peak", "piq_timing_peak", "y_timing_peak",
    "irate_timing_peak", "rrate_timing_peak", "pi_timing_peak", "piq_timing_theorypeak",
    "piq_value_theorypeak", "y_timing_theorypeak", "y_value_theorypeak",
    "irate_timing_theorypeak", "irate_value_theorypeak", "rrate_timing_theorypeak",
    "rrate_value_theorypeak", "pi_timing_theorypeak", "pi_value_theorypeak",
    "piq_timing_agnpeak", "piq_value_agnpeak", "y_timing_agnpeak", "y_value_agnpeak",
    "irate_timing_agnpeak", "irate_value_agnpeak", "rrate_timing_agnpeak",
    "rrate_value_agnpeak", "pi_timing_agnpeak", "pi_value_agnpeak", "IScurve20",
    "infl_per_rr20", "Billsacrat20", "IScurve40", "infl_per_rr40", "Billsacrat40",
    "IScurve60", "infl_per_rr60", "Billsacrat60", "stky_pr_nondx", "stky_wg_nondx",
    "stky_all", "stky_pr_wg_ndx", "ndx_all", "ln_neq",
]
for col in legacy_reg_float32_columns:
    reg[col] = reg[col].astype(np.float32)
for horizon in horizons:
    reg[f"IScurve{horizon}"] = (reg[f"y_cum{horizon}"] / reg[f"rrate_cum{horizon}"]).astype(np.float32)
    reg[f"infl_per_rr{horizon}"] = (reg[f"piq_cum{horizon}"] / reg[f"rrate_cum{horizon}"]).astype(np.float32)
    reg[f"Billsacrat{horizon}"] = (reg[f"piq_cum{horizon}"] / reg[f"y_cum{horizon}"]).astype(np.float32)

reg.to_stata(output_dir / "MMB_reg_format.dta", write_index=False, version=118)
reg.to_excel(output_dir / "MMB_reg_format.xlsx", index=False)


# Step 6: Build the cloud-graph IRF panel with model attributes and VAR benchmark.
irf_full = irf.loc[irf["rule"] != "Model"].copy()
irf_full = irf_full.drop(columns=["id"])
irf_full["period"] = irf_full["period"] - 1
irf_full = irf_full.loc[irf_full["period"] <= cloud_graph_extent].copy()
irf_full["id"] = irf_full.groupby(["model", "rule"], sort=True).ngroup() + 1
irf_full = irf_full.merge(rhs, on="model", how="left")
irf_full["_merge"] = 3
irf_full["model_type_n"] = np.nan
irf_full.loc[irf_full["calibrated"] == 1, "model_type_n"] = 1
irf_full.loc[irf_full["estimated"] == 1, "model_type_n"] = 2
irf_full.loc[(irf_full["calibrated"] == 1) & (irf_full["estimated"] == 1), "model_type_n"] = 3
irf_full["model_type"] = ""
irf_full.loc[irf_full["calibrated"] == 1, "model_type"] = "Calibrated"
irf_full.loc[irf_full["estimated"] == 1, "model_type"] = "Estimated"
irf_full.loc[(irf_full["calibrated"] == 1) & (irf_full["estimated"] == 1), "model_type"] = "Both"

bob = pd.read_csv(input_dir / "bob_var_irfs.csv")
bob.columns = [c.lower() for c in bob.columns]
bob = bob.rename(columns={"rtb": "irate", "pi": "piq"})
bob["model"] = "VAR, 1963:Q1-2007:Q4"
bob["rule"] = "NA"
bob["period"] = np.arange(len(bob))
for col in ["irate", "piq", "y"]:
    bob[col] = (-pd.to_numeric(bob[col], errors="coerce")).astype(np.float32).astype(float)
bob["rrate"] = (bob["irate"].astype(np.float32) - bob["piq"].shift(-1).astype(np.float32)).astype(np.float32).astype(float)
bob = bob.loc[bob["period"] <= cloud_graph_extent, ["period", "model", "rule", "piq", "y", "irate", "rrate"]].copy()
bob["pi"] = np.nan
bob["id"] = np.nan

for col in irf_full.columns:
    if col not in bob.columns:
        bob[col] = np.nan
bob = bob[irf_full.columns]
irf_full = pd.concat([irf_full, bob], ignore_index=True)
irf_full = irf_full.drop(columns=[c for c in ["cites", "cites_w1", "cites_w2", "cites_w3"] if c in irf_full.columns])
irf_full = irf_full.sort_values(["model", "rule", "period"]).reset_index(drop=True)
legacy_irf_full_columns = [
    "id", "period", "model", "rule", "piq", "y", "irate", "rrate", "pi",
    "cb_authors", "estimated", "calibrated", "neq", "open", "firm_bs", "bank",
    "hh_demand", "labor_frict", "gov_spend", "tax", "gov_debt", "learning", "stky_wg",
    "pr_ndx", "wg_ndx", "est_early", "est_late", "vint_early", "vint_mid",
    "vint_late", "pub_date", "est_start", "est_end", "stky_pr_calvo",
    "stky_pr_rotemberg", "stky_pr_other", "not_stky_pr", "stky_pr",
    "wg_ndx_mult", "wg_ndx_prprice", "wg_ndx_other", "stky_pr_ndx",
    "stky_pr_nondx", "stky_wg_ndx", "stky_wg_nondx", "stky_all",
    "stky_pr_wg_ndx", "ndx_all", "other_channel", "fiscal", "ln_neq",
    "cb_authors_ext", "_merge", "model_type_n", "model_type",
]
irf_full = irf_full[legacy_irf_full_columns]
legacy_irf_full_float32_columns = [
    "id", "period", "est_early", "est_late", "vint_early", "vint_mid",
    "vint_late", "stky_pr_calvo", "stky_pr_rotemberg", "stky_pr_other",
    "not_stky_pr", "stky_pr", "wg_ndx_mult", "wg_ndx_prprice", "wg_ndx_other",
    "stky_pr_ndx", "stky_pr_nondx", "stky_wg_ndx", "stky_wg_nondx",
    "stky_all", "stky_pr_wg_ndx", "ndx_all", "other_channel", "fiscal",
    "ln_neq", "cb_authors_ext", "model_type_n",
]
for col in legacy_irf_full_float32_columns:
    irf_full[col] = irf_full[col].astype(np.float32)
irf_full.to_stata(output_dir / "MMB_IRF_format_full.dta", write_index=False, version=118)


# Step 7: Write text metadata for the stable binary outputs.
write_codebook(
    irf,
    output_dir / "MMB_IRF_format_codebook.txt",
    "MMB response CSVs imported through tasks/data/import_mmb_legacy_data",
    "Dropped duplicate, non-US calibrated, and excluded models listed in config/params.yaml; dropped model-rule responses; sign-flipped responses; period zero removed.",
)
write_codebook(
    irf_full,
    output_dir / "MMB_IRF_format_full_codebook.txt",
    "MMB_IRF_format.dta plus hand-coded model attributes and Bob VAR benchmark",
    f"Restricted to periods 0 through {cloud_graph_extent} after aligning the shock to period 0 for cloud graphs.",
)
write_codebook(
    reg,
    output_dir / "MMB_reg_format_codebook.txt",
    "Constructed IRF outcomes, sacrifice ratios, and hand-coded model attributes",
    "Merged horizons, sacrifice ratios, and RHS variables; model exclusions from config/params.yaml applied.",
)

parity_sections = ["Analysis dataset parity report", "==============================", ""]
legacy_derived_dir = repo_dir / "legacy" / "mmb_upgraded" / "data" / "derived"
parity_specs = [
    ("MMB_IRF_format", irf, legacy_derived_dir / "MMB_IRF_format.dta", ["model", "rule", "period"]),
    ("lhs_20", lhs_by_horizon[20], legacy_derived_dir / "lhs_20.dta", ["model", "rule"]),
    ("lhs_40", lhs_by_horizon[40], legacy_derived_dir / "lhs_40.dta", ["model", "rule"]),
    ("lhs_60", lhs_by_horizon[60], legacy_derived_dir / "lhs_60.dta", ["model", "rule"]),
    ("sacratios_20", sacratio_by_horizon[20], legacy_derived_dir / "sacratios_20.dta", ["model", "rule"]),
    ("sacratios_40", sacratio_by_horizon[40], legacy_derived_dir / "sacratios_40.dta", ["model", "rule"]),
    ("sacratios_60", sacratio_by_horizon[60], legacy_derived_dir / "sacratios_60.dta", ["model", "rule"]),
    ("rhs", rhs, legacy_derived_dir / "rhs.dta", ["model"]),
    ("MMB_reg_format", reg, legacy_derived_dir / "MMB_reg_format.dta", ["model", "rule"]),
    ("MMB_IRF_format_full", irf_full, legacy_derived_dir / "MMB_IRF_format_full.dta", ["model", "rule", "period"]),
]
for name, rebuilt, legacy_path, key_columns in parity_specs:
    parity_sections.extend(compare_for_report(name, rebuilt, legacy_path, key_columns))
    parity_sections.append("")

with open(output_dir / "parity_report.txt", "w", encoding="utf-8") as f:
    f.write("\n".join(parity_sections))

print(f"Wrote IRF panel with {irf.shape[0]} rows and {irf.shape[1]} columns.")
print(f"Wrote cloud-graph IRF panel with {irf_full.shape[0]} rows and {irf_full.shape[1]} columns.")
print(f"Wrote regression panel with {reg.shape[0]} rows and {reg.shape[1]} columns.")
