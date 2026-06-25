"""
Purpose:
    Apply Gemini discrepancy corrections to the model-characteristics workbook.

Inputs:
    ../input/Model_Characteristics_corrections.xlsx
    ../output/model_audit.csv

Outputs:
    ../output/Model_Characteristics_corrections_llm.xlsx
    ../output/Model_Characteristics_corrections_llm_summary.txt

Run:
    make corrections-workbook from tasks/data/classifying_v2/code/
"""

from pathlib import Path
import csv
import re

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font


code_dir = Path(__file__).resolve().parent
task_dir = code_dir.parent
input_dir = task_dir / "input"
output_dir = task_dir / "output"

source_workbook_path = input_dir / "Model_Characteristics_corrections.xlsx"
audit_path = output_dir / "model_audit.csv"
output_workbook_path = output_dir / "Model_Characteristics_corrections_llm.xlsx"
summary_path = output_dir / "Model_Characteristics_corrections_llm_summary.txt"

first_model_row = 2
last_model_row = 93

percent_variables = {"CB_Authors"}
boolean_variables = {
    "Working_Paper",
    "Published",
    "Calibrated",
    "Open",
    "Gov_Spend",
    "Tax",
    "Gov_Debt",
    "Learning",
    "Rational_Expectations",
    "Lagged_Terms",
    "Firm_Balance_Sheet_Channel",
    "Bank_Lending_Intermediation_Channel",
    "Constrained_Household_Demand_Channel",
    "Real_Labor_Market_Friction",
    "Sticky_Prices",
    "Sticky_Wages",
    "Price_Indexation",
    "Wage_Indexation",
}
estimate_range_variables = {"Est_Date_Range_Start", "Est_Date_Range_End"}
date_variables = {"Date_Pub"}

# These audit rows are valid even though the current workbook has no columns for them.
# They are preserved in the log sheet, and will be written normally if columns are added.
classification_only_variables = {
    "Firm_Balance_Sheet_Channel",
    "Bank_Lending_Intermediation_Channel",
    "Constrained_Household_Demand_Channel",
    "Real_Labor_Market_Friction",
}

category_spelling = {
    "all sectors": "Everywhere",
    "intermediate goods firms": "Intermediate Firms",
    "final goods firms": "Final Goods Firms",
    "prev price inflation": "Prev Price Inflation",
    "prev wage inflation": "Prev Wage Inflation",
    "prev wages": "Prev Wages",
    "steady state inflation": "Steady-State Inflation",
    "steady-state inflation": "Steady-State Inflation",
    "multiple": "Multiple",
    "other": "Other",
    "partial": "Partial",
    "full": "Full",
    "multiple sectors": "Multiple Sectors",
    "representative single sector": "Representative Single Sector",
    "final/retail firms": "Final/Retail Firms",
    "final retail firms": "Final/Retail Firms",
    "calvo": "Calvo",
    "rotemberg": "Rotemberg",
    "wage contracting": "Wage Contracting",
    "bargaining": "Bargaining",
}

month_names = (
    "January|February|March|April|May|June|July|August|September|October|November|December|"
    "Jan|Feb|Mar|Apr|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec"
)
roman_quarters = {"i": "1", "ii": "2", "iii": "3", "iv": "4"}


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_boolean(value):
    text = clean_text(value).lower()
    if text in {"true", "yes", "1"}:
        return True
    if text in {"false", "no", "0", "no mention"}:
        return False
    raise SystemExit(f"Could not normalize boolean correction: {value}")


def normalize_estimated_status(value):
    text = clean_text(value).lower()
    if text == "estimated":
        return True, False
    if text == "calibrated":
        return False, True
    raise SystemExit(f"Could not normalize estimated/calibrated correction: {value}")


def normalize_percent(value):
    text = clean_text(value)
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        raise SystemExit(f"Could not normalize percentage correction: {value}")
    return round(float(match.group(0)) / 100, 6)


def normalize_estimate_range(value):
    text = clean_text(value)
    compact = text.replace(" ", "")

    match = re.search(r"^(\d{4})[Qq]([1-4])$", compact)
    if match:
        return f"{match.group(1)}Q{match.group(2)}"

    match = re.search(r"^(\d{4})[Qq]([IVXivx]+)$", compact)
    if match:
        roman = roman_quarters.get(match.group(2).lower())
        if roman:
            return f"{match.group(1)}Q{roman}"

    match = re.search(r"^(\d{4}):[Qq]?([1-4])$", compact)
    if match:
        return f"{match.group(1)}Q{match.group(2)}"

    match = re.search(r"^(\d{4}):([IVXivx]+)$", compact)
    if match:
        roman = roman_quarters.get(match.group(2).lower())
        if roman:
            return f"{match.group(1)}Q{roman}"

    match = re.search(r"^(\d{4})$", compact)
    if match:
        return match.group(1)

    return text


def normalize_publication_date(value):
    text = clean_text(value)

    # Keep the journal date when Gemini returned both working-paper and journal dates.
    journal_match = re.search(
        rf"(({month_names})\s+(?:\d{{1,2}},\s*)?\d{{4}})\s*\(?journal",
        text,
        flags=re.IGNORECASE,
    )
    if journal_match:
        return journal_match.group(1)

    leading_year = re.search(r"^\s*((?:19|20)\d{2})\b", text)
    if leading_year:
        return leading_year.group(1)

    month_day_year = re.search(rf"({month_names})\s+\d{{1,2}},\s*\d{{4}}", text, flags=re.IGNORECASE)
    if month_day_year:
        return month_day_year.group(0)

    month_year = re.search(rf"({month_names})\s+\d{{4}}", text, flags=re.IGNORECASE)
    if month_year:
        return month_year.group(0)

    year = re.search(r"\b(19|20)\d{2}\b", text)
    if year:
        return year.group(0)

    return text


def normalize_category(value):
    text = clean_text(value)
    normalized = category_spelling.get(text.lower())
    if normalized:
        return normalized
    return text


output_dir.mkdir(parents=True, exist_ok=True)

if not source_workbook_path.exists():
    raise SystemExit(f"Missing source workbook: {source_workbook_path}")
if not audit_path.exists():
    raise SystemExit(f"Missing audit file: {audit_path}")

# Step 1: Load the workbook exactly once and identify the data surface downstream reads.
workbook = load_workbook(source_workbook_path)
main_sheet = workbook[workbook.sheetnames[0]]
header_by_column = {col: clean_text(main_sheet.cell(1, col).value) for col in range(1, main_sheet.max_column + 1)}
column_by_header = {header: col for col, header in header_by_column.items() if header}

model_rows = {}
for row in range(first_model_row, last_model_row + 1):
    model = clean_text(main_sheet.cell(row, column_by_header["Model"]).value)
    if not model:
        raise SystemExit(f"Blank model in first 92-row analysis range at row {row}")
    if model in model_rows:
        raise SystemExit(f"Duplicate model in first 92-row analysis range: {model}")
    model_rows[model] = row

# Step 2: Read and validate the Gemini discrepancy rows before changing any cells.
audit_rows = []
classification_only_rows = []
seen_pairs = set()
with open(audit_path, "r", encoding="utf-8", newline="") as f:
    reader = csv.DictReader(f)
    for row in reader:
        model = clean_text(row.get("model"))
        variable = clean_text(row.get("variable"))
        right_coding = clean_text(row.get("right_coding"))
        explanation = clean_text(row.get("explanation"))
        if not model or not variable:
            raise SystemExit(f"Malformed audit row: {row}")
        pair = (model, variable)
        if pair in seen_pairs:
            raise SystemExit(f"Duplicate audit correction: {model} {variable}")
        if model not in model_rows:
            raise SystemExit(f"Audit correction model not in first 92 workbook rows: {model}")
        if variable not in column_by_header:
            if variable not in classification_only_variables:
                raise SystemExit(f"Audit correction variable not in workbook headers: {variable}")
            seen_pairs.add(pair)
            classification_only_rows.append(
                {
                    "model": model,
                    "variable": variable,
                    "right_coding": right_coding,
                    "explanation": explanation,
                }
            )
            continue
        seen_pairs.add(pair)
        audit_rows.append(
            {
                "model": model,
                "variable": variable,
                "right_coding": right_coding,
                "explanation": explanation,
            }
        )

# Step 3: Apply normalized values to the first sheet, preserving the original workbook columns.
applied_rows = []
for correction in audit_rows:
    model = correction["model"]
    variable = correction["variable"]
    raw_right_coding = correction["right_coding"]
    row_number = model_rows[model]

    if variable == "Estimated":
        estimated_value, calibrated_value = normalize_estimated_status(raw_right_coding)
        for paired_variable, corrected_value in [
            ("Estimated", estimated_value),
            ("Calibrated", calibrated_value),
        ]:
            column_number = column_by_header[paired_variable]
            cell = main_sheet.cell(row_number, column_number)
            original_value = cell.value
            cell.value = corrected_value
            cell.comment = Comment(
                "LLM correction from tasks/data/classifying_v2/output/model_audit.csv\n"
                f"Original value: {clean_text(original_value)}\n"
                f"LLM right_coding for Estimated/Calibrated: {raw_right_coding}\n"
                f"Explanation: {correction['explanation']}",
                "Codex",
            )
            applied_rows.append(
                [
                    model,
                    paired_variable,
                    original_value,
                    corrected_value,
                    raw_right_coding,
                    correction["explanation"],
                ]
            )
        continue

    column_number = column_by_header[variable]
    cell = main_sheet.cell(row_number, column_number)
    original_value = cell.value

    if variable in percent_variables:
        percent_source = raw_right_coding
        if not re.search(r"-?\d+(?:\.\d+)?", percent_source):
            percent_source = f"{raw_right_coding} {correction['explanation']}"
        corrected_value = normalize_percent(percent_source)
        cell.number_format = "0.00%"
    elif variable in boolean_variables:
        corrected_value = normalize_boolean(raw_right_coding)
    elif variable in estimate_range_variables:
        corrected_value = normalize_estimate_range(raw_right_coding)
    elif variable in date_variables:
        corrected_value = normalize_publication_date(raw_right_coding)
    else:
        corrected_value = normalize_category(raw_right_coding)

    cell.value = corrected_value
    cell.comment = Comment(
        "LLM correction from tasks/data/classifying_v2/output/model_audit.csv\n"
        f"Original value: {clean_text(original_value)}\n"
        f"LLM right_coding: {raw_right_coding}\n"
        f"Explanation: {correction['explanation']}",
        "Codex",
    )

    applied_rows.append(
        [
            model,
            variable,
            original_value,
            corrected_value,
            raw_right_coding,
            correction["explanation"],
        ]
    )

# Step 4: Add a readable audit trail sheet without changing the first sheet used by pandas.
if "LLM_Corrections_Log" in workbook.sheetnames:
    del workbook["LLM_Corrections_Log"]
log_sheet = workbook.create_sheet("LLM_Corrections_Log")
log_headers = [
    "Model",
    "Variable",
    "Original_Value",
    "Value_Written",
    "LLM_Right_Coding",
    "LLM_Explanation",
]
log_sheet.append(log_headers)
for applied in applied_rows:
    log_sheet.append(applied)
for correction in classification_only_rows:
    log_sheet.append(
        [
            correction["model"],
            correction["variable"],
            "",
            "Not written: classification-only variable absent from workbook",
            correction["right_coding"],
            correction["explanation"],
        ]
    )

for col in range(1, len(log_headers) + 1):
    log_sheet.cell(1, col).font = Font(bold=True)
log_sheet.freeze_panes = "A2"
log_sheet.auto_filter.ref = log_sheet.dimensions
log_sheet.column_dimensions["A"].width = 16
log_sheet.column_dimensions["B"].width = 26
log_sheet.column_dimensions["C"].width = 22
log_sheet.column_dimensions["D"].width = 22
log_sheet.column_dimensions["E"].width = 28
log_sheet.column_dimensions["F"].width = 90

workbook.save(output_workbook_path)

with open(summary_path, "w", encoding="utf-8") as f:
    f.write("Model_Characteristics_corrections_llm workbook summary\n")
    f.write("=====================================================\n\n")
    f.write(f"Source workbook: {source_workbook_path.relative_to(task_dir)}\n")
    f.write(f"Audit file: {audit_path.relative_to(task_dir)}\n")
    f.write(f"Output workbook: {output_workbook_path.relative_to(task_dir)}\n")
    f.write(f"Model rows eligible for downstream analysis: {len(model_rows)}\n")
    f.write(f"LLM corrections applied: {len(applied_rows)}\n")
    f.write(f"Classification-only rows logged but not written: {len(classification_only_rows)}\n")
    f.write("The first sheet preserves the original workbook layout and applies corrected values in-place.\n")
    f.write("The LLM_Corrections_Log sheet records original values, normalized values, raw LLM codings, explanations, and classification-only rows absent from the workbook.\n")
